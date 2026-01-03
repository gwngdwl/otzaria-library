import csv
from collections import defaultdict
from collections.abc import Generator
from pathlib import Path
from time import sleep

import mwparserfromhell
import requests
from openpyxl import load_workbook

API_URL = "https://he.wikisource.org/w/api.php"
BASE_URL = "https://he.wikisource.org/wiki/"
API_URL = "https://wiki.jewishbooks.org.il/mediawiki/api.php"
BASE_URL = "https://wiki.jewishbooks.org.il/mediawiki/wiki/"
SESSION = requests.Session()
HEADERS = {
    'User-Agent': 'OtzariaLibraryBot/1.0 (https://otzaria.com; bot for transferring content from Wikisource to Otzaria)',
}


def get_old_data(csv_file_path: Path) -> defaultdict[str, defaultdict[str, set]]:
    data = defaultdict(lambda: defaultdict(set))
    if not csv_file_path.exists():
        return data
    with csv_file_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            data[row[0]]["param_len"] = set(map(int, row[2].split(","))) if row[2] else set()
            data[row[0]]["books"] = set(row[3].split(","))
            data[row[0]]["action"] = set(row[4])
            data[row[0]]["action_comments"] = set(row[5])
            data[row[0]]["template_desc"] = set(row[6])
            data[row[0]]["comments"] = set(row[7])
    return data


def get_page_content(page_title: str) -> str:
    params = {
        'action': 'query',
        'prop': 'revisions',
        'titles': page_title,
        'rvslots': '*',
        'rvprop': 'content',
        'format': 'json',
    }
    try:
        response = SESSION.get(API_URL, params=params, headers=HEADERS)
        response.raise_for_status()

        data = response.json()

        # בדיקת התוכן שהתקבל
        pages = data.get("query", {}).get("pages", {})
        for page_id, page_data in pages.items():
            if "revisions" in page_data:
                return page_data["revisions"][0]["slots"]["main"]["*"]
        return "Page not found or no content available."

    except requests.exceptions.RequestException as e:
        print(f"Error fetching page content: {e}")
        raise


def filter_templates(string: str) -> Generator[tuple[str, int], None, None]:
    parsed: mwparserfromhell.wikicode.Wikicode = mwparserfromhell.parse(string)
    templates = parsed.filter_templates(recursive=True)
    if templates:
        for template in templates:
            template_name = str(template.name).strip()
            if template_name.startswith("#") and ":" in template_name:
                template_name = template_name.split(":", 1)[0].strip()
            yield template_name, len(template.params)


def iter_books_with_author(xlsx_file_path: Path) -> Generator[tuple[str, list[list[str]]]]:
    wb = load_workbook(xlsx_file_path, data_only=True, read_only=True)
    if "ספרים" not in wb.sheetnames:
        raise ValueError("אין גיליון בשם 'ספרים' בקובץ")

    books_ws = wb["ספרים"]

    name_idx = 0
    done_idx = 2

    for row in books_ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        book_name = row[name_idx]
        done = row[done_idx]
        if not book_name or book_name not in wb.sheetnames:
            continue
        # if done is True:
        #     continue

        ws = wb[book_name]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # דלג על שורת כותרת
        yield book_name, ([["" if c is None else str(c) for c in r] for r in rows_iter])

    wb.close()


def main():
    output_file = Path(__file__).parent / "templates_2.csv"
    books_templates = get_old_data(output_file)
    # for book_name, rows in iter_books_with_author(Path(__file__).parent / "ויקיטקסט_4.xlsx"):
    for book_name, rows in iter_books_with_author(Path(r"C:\Users\User\Downloads\אוצר הספרים היהודי השיתופי_4.xlsx")):
        print(f"ספר: {book_name}, דפים: {len(rows)}")
        for row in rows:
            content = get_page_content(row[0])
            sleep(1)  # הוספת השהייה של שנייה בין בקשות
            for template_name, param_count in filter_templates(content):
                books_templates[template_name]["param_len"].add(param_count)
                books_templates[template_name]["books"].add(book_name)
    with output_file.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["תבנית", "קישור לדף התבנית", "מספר פרמטרים", "ספרים", "פעולה לביצוע", "הערות לביצוע", "הסבר על התבנית", "הערות כלליות"])

        for template_name in sorted(books_templates):
            template_link = f"{BASE_URL}תבנית:{template_name.replace(' ', '_')}"
            writer.writerow(
                [
                    template_name,
                    template_link,
                    ",".join(map(str, sorted(books_templates[template_name]["param_len"]))),
                    ",".join(sorted(books_templates[template_name]["books"])),
                    next(iter(books_templates[template_name]["action"])) if books_templates[template_name]["action"] else "",
                    next(iter(books_templates[template_name]["action_comments"])) if books_templates[template_name]["action_comments"] else "",
                    next(iter(books_templates[template_name]["template_desc"])) if books_templates[template_name]["template_desc"] else "",
                    next(iter(books_templates[template_name]["comments"])) if books_templates[template_name]["comments"] else "",
                ]
            )


if __name__ == "__main__":
    main()
