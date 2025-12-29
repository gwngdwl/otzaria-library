import csv
import html
import json
import re
import time
from collections.abc import Generator
from pathlib import Path

from mediawikitootzaria import htmltootzaria, mediawikiapi, mediawikitohtml, templates, utils
from openpyxl import load_workbook
from tqdm import tqdm

mediawikiapi.BASE_URL = mediawikiapi.WIKISOURCE
templates.replacement_dict = templates.wikisource_replacement_dict


def iter_books_with_author(xlsx_file_path: Path) -> Generator[tuple[str, str, str, list[list[str]]]]:
    wb = load_workbook(xlsx_file_path, data_only=True, read_only=True)
    if "ספרים" not in wb.sheetnames:
        raise ValueError("אין גיליון בשם 'ספרים' בקובץ")

    books_ws = wb["ספרים"]

    name_idx = 0
    author_idx = 1
    done_idx = 2
    link_idx = 3

    for row in books_ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        book_name = row[name_idx]
        author = row[author_idx]
        done = row[done_idx]
        link = row[link_idx]
        if not book_name or book_name not in wb.sheetnames:
            continue
        if done is True:
            continue

        ws = wb[book_name]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # דלג על שורת כותרת
        yield author, book_name, link, ([["" if c is None else str(c) for c in r] for r in rows_iter])

    wb.close()


def read_order_from_csv(csv_file_path: Path) -> Generator[list[str]]:
    with csv_file_path.open("r", encoding="utf-8") as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader)
        yield from csv_reader


# def main(csv_file_path: Path, title: str, author: str) -> None:
def process_book(pages: list[list[str]], title: str, author: str, link: str) -> None:
    file_name = utils.sanitize_filename(title)
    target_dir = Path("ספרים")
    target_dir.mkdir(exist_ok=True, parents=True)
    sup_num = 0
    all_sup = []
    dict_links = []
    # all_content = [f"<h1>{title}</h1>", author if author else "", "באדיבות 'אוצר הספרים היהודי השיתופי'", f'(במצב מקוון אפשר ללחוץ <a href="{link}">כאן</a> ולתקן את הדף המקורי או להוסיף הערת שוליים)']
    all_content = [f"<h1>{title}</h1>", author if author else ""]
    lines = len(all_content)
    # for book in read_order_from_csv(csv_file_path):
    time.sleep(1)
    h_level = 2
    for page in tqdm(pages):
        page_url = page[0]
        content = mediawikiapi.get_page_content(page_url)
        for index, i in enumerate(page[3:], start=2):
            if i:
                all_content.append(f"<h{index}>{i}</h{index}>")
                lines += 1
                h_level = index + 1

        content = mediawikitohtml.media_wiki_list_to_html(content)
        content = mediawikitohtml.wikitext_to_html(content, h_level)
        content, sup = templates.remove_templates(content)
        content = mediawikitohtml.fix_new_lines(content)
        content = htmltootzaria.process_body_html(content, h_level)
        content = htmltootzaria.adjust_html_tag_spaces(content)
        content = re.sub(r"<קטע (?:התחלה|סוף)=[^>]+/?>", "", content)
        strip_all = content.split("\n")
        all_content.extend(strip_all)
        for index, line in enumerate(strip_all, start=lines + 1):
            find = re.findall(r'<sup style="color: gray;">(\d+)</sup>', line)
            dict_links.extend([
                {
                    "line_index_1": index,
                    "heRef_2": "הערות",
                    "path_2": f"הערות על {file_name}.txt",
                    "line_index_2": int(i) + sup_num,
                    "Conection Type": "commentary"
                } for i in find
            ])
            # for i in find:
            #     dict_links.append({
            #         "line_index_1": index,
            #         "heRef_2": "הערות",
            #         "path_2": f"הערות על {title}.txt",
            #         "line_index_2": int(i) + sup_num,
            #         "Conection Type": "commentary"
            #     })

        lines += len(strip_all)

        for k, value in sup.items():
            v = mediawikitohtml.wikitext_to_html(value)
            v = htmltootzaria.fix_comments(v)
            v = html.unescape(v)
            all_sup.append(f'{k} {v}')
            sup_num += 1
    file_path = target_dir / f"{file_name}.txt"
    json_file_path = target_dir / f"{file_name}_links.json"
    comments_file_path = target_dir / f"הערות על {file_name}.txt"

    if dict_links:
        with json_file_path.open('w', encoding='utf-8') as f:
            json.dump(dict_links, f)
    with file_path.open("w", encoding="utf-8") as f:
        f.write("\n".join(all_content))
    if all_sup:
        with comments_file_path.open("w", encoding="utf-8") as f:
            f.write("\n".join(all_sup))


# file_path = r"C:\Users\User\Downloads\אוצר הספרים היהודי השיתופי - עץ הדר.csv"
# book_title = "עץ הדר"
# book_author = ""
def main() -> None:
    for author, book_name, link, rows in iter_books_with_author(Path(r"C:\Users\User\Downloads\ויקיטקסט_5.xlsx")):
        print("מחבר:", author)
        print(f"{book_name=}:")
        # main(Path(file_path), book_title, book_author)
        process_book(rows, book_name, author, link)
        print(f"סיים: {book_name}")


if __name__ == "__main__":
    main()
